####################################################################
# Name：Achievement management system                              #
# Author：Guan Ruiqi                                               #
# Date：April 20, 2020                                             #
# Function description：Student achievement self management system #
# Ver：V 1.00                                                      #
####################################################################

import matplotlib.pyplot as plt
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import datetime

#显示中文
plt.rcParams['font.sans-serif'] = ['SimHei']
#记录显示的曲线
score_line = [None,None,None,None,None,None]

#语文成绩变化曲线
def score_yuwen( ):
    global score_line,img_src,my_image
    Wb = openpyxl.load_workbook("语文.xlsx")
    sheet = Wb['成绩单']
    date = []
    score = []
    scoreA = []
    linenum = sheet.max_row
    for i in range(2,(linenum+1)):
        date.append(sheet["A" + str(i)].value)
        score.append(sheet["B" + str(i)].value)
        scoreA.append(sheet["C" + str(i)].value)
    date = str(date)
    plt.clf()
    score_line[0]=plt.plot(score,'r-d', linewidth=2)
    score_line[1]=plt.plot(scoreA,'r-.*', linewidth=1)
    plt.title("语文")
    plt.savefig("成绩.png",)
    img_src = tk.PhotoImage(file = "成绩.png")
    my_canvas.itemconfig(my_image,image=img_src,anchor="nw")
    my_canvas.update()

#数学成绩变化曲线
def score_math( ):
    global score_line,img_src,my_image
    Wb = openpyxl.load_workbook("数学.xlsx")
    sheet = Wb['成绩单']
    date = []
    score = []
    scoreA = []
    linenum = sheet.max_row
    for i in range(2,(linenum+1)):
        date.append(sheet["A" + str(i)].value)
        score.append(sheet["B" + str(i)].value)
        scoreA.append(sheet["C" + str(i)].value)
    date = str(date)
    plt.clf()
    score_line[0]=plt.plot(score,'b-d', linewidth=2)
    score_line[1]=plt.plot(scoreA,'b-.*', linewidth=1)
    plt.title("数学")
    plt.savefig("成绩.png",)
    img_src = tk.PhotoImage(file = "成绩.png")
    my_canvas.itemconfig(my_image,image=img_src,anchor="nw")
    my_canvas.update()

#英语成绩变化曲线
def score_eng( ):
    global score_line,img_src,my_image
    Wb = openpyxl.load_workbook("英语.xlsx")
    sheet = Wb['成绩单']
    date = []
    score = []
    scoreA = []
    linenum = sheet.max_row
    for i in range(2,(linenum+1)):
        date.append(sheet["A" + str(i)].value)
        score.append(sheet["B" + str(i)].value)
        scoreA.append(sheet["C" + str(i)].value)
    date = str(date)
    plt.clf()
    score_line[0]=plt.plot(score,'y-d', linewidth=2)
    score_line[1]=plt.plot(scoreA,'y-.*', linewidth=1)
    plt.title("英语")
    plt.savefig("成绩.png",)
    img_src = tk.PhotoImage(file = "成绩.png")
    my_canvas.itemconfig(my_image,image=img_src,anchor="nw")
    my_canvas.update()


def update_excel( Name, date, score, scoreA):
    Wb = openpyxl.load_workbook(Name + ".xlsx")
    sheet = Wb['成绩单']
    linenum = sheet.max_row
    #更新成绩单
    sheet.cell(row = linenum+1, column = 1).value = date
    sheet.cell(row = linenum+1, column = 2).value = score
    sheet.cell(row = linenum+1, column = 3).value = scoreA
    Wb.save(Name + ".xlsx")
    tk.messagebox.showinfo('','成绩更新成功！')

def scoreupdate():
    global date,score,scoreA
    #获取输入框内容
    dateget = date.get()
    scoreget = score.get()
    scoreAget = scoreA.get()
    numberi = number.get()
    numberi = str(numberi)

    if dateget == 'none' or scoreget == 'none' or scoreAget == 'none':
        tk.messagebox.showerror('ERROR','您没有输入数值!')
    else:
         update_excel( numberi, dateget, int(scoreget), int(scoreAget) )
    date.set(dateget)
    score.set(scoreget)
    scoreA.set(scoreAget)

def score_clear():
    global img_src,my_image
    img_src = tk.PhotoImage(file = "bank.png")
    my_canvas.itemconfig(my_image,image=img_src,anchor="nw")
    my_canvas.update()

    
#生成窗口
window = tk.Tk()
window.title('成绩分析')
window.maxsize(750,700)
window.minsize(750,700)

#生成客户选择菜单
classlabel = ttk.Label(window , text="请选择科目")
classlabel.place(x = 50 , y = 20)
number = tk.StringVar()
numberChosen = ttk.Combobox(window , width=16, textvariable=number, state='readonly')
numberChosen['values'] = ("语文", "数学", "英语")
numberChosen.place(x = 50 , y = 40 )
numberChosen.current(0)

classlabel = ttk.Label(window , text="请输入考试日期")
classlabel.place(x = 50 , y = 80)
date = tk.StringVar()
datatoday = datetime.datetime.now()
date.set(str(datatoday.year)+"年"+str(datatoday.month)+"月"+str(datatoday.day)+"日")
dateenter = ttk.Entry(window , width=16, textvariable=date )
dateenter.place(x = 50 , y = 100 )

#输入考试成绩对话框
classlabel = ttk.Label(window , text="请输入考试成绩")
classlabel.place(x = 220 , y = 80)
score = tk.StringVar()
score.set('none')
scoreenter = ttk.Entry(window , width = 16, textvariable = score)
scoreenter.place(x = 220 , y = 100)

#输入班级平均成绩对话框
classlabel = ttk.Label(window , text="请输入班级平均成绩")
classlabel.place(x = 390 , y = 80)
scoreA = tk.StringVar()
scoreA.set("none")
scoreAenter = ttk.Entry(window , width = 16, textvariable = scoreA)
scoreAenter.place(x = 390 , y = 100)

classlabel = ttk.Label(window , text="确定提交成绩")
classlabel.place(x = 550 , y = 80)
action = ttk.Button(window , width=16, text="提交成绩", command=scoreupdate)
action.place(x = 550 , y = 100)

classlabel = ttk.Label(window , text="成绩变化趋势")
classlabel.place(x = 50 , y = 140)

action = ttk.Button(window , width=16, text="语文成绩", command=score_yuwen)
action.place(x = 50 , y = 160)

#数学成绩显示按钮
action = ttk.Button(window , width=16, text="数学成绩", command=score_math)
action.place(x = 220 , y = 160)

#英语成绩显示按钮
action = ttk.Button(window , width=16, text="英语成绩", command=score_eng)
action.place(x = 390 , y = 160)

#清除曲线
action = ttk.Button(window , width=16, text="清除曲线", command=score_clear)
action.place(x = 550 , y = 160)

#生成显示成绩曲线的画板，画板名称window,默认图片"bank.png"
my_canvas = tk.Canvas(window , width = 640 , height = 480 , bg = "white")
my_canvas.place(x = 50 , y = 200)
img_src = tk.PhotoImage(file = "bank.png")
my_image = my_canvas.create_image(0 , 0 , image = img_src , anchor = "nw")

window.mainloop()